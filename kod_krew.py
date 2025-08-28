import glob
import os
import shutil
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox

class RaportApp:
    def __init__(self, master):
        self.master = master
        master.title("Generator raportów z wyników krwi")
        master.geometry("500x300")

        self.folder_in = None
        self.template_path = None
        self.folder_out = None

        # Przyciski i pola tekstowe
        tk.Button(master, text="Wybierz folder wejściowy (XML)", command=self.choose_input).pack(pady=5)
        self.label_in = tk.Label(master, text="Nie wybrano folderu")
        self.label_in.pack()

        tk.Button(master, text="Wybierz plik szablonu (XLSX)", command=self.choose_template).pack(pady=5)
        self.label_template = tk.Label(master, text="Nie wybrano pliku")
        self.label_template.pack()

        tk.Button(master, text="Wybierz folder wyjściowy", command=self.choose_output).pack(pady=5)
        self.label_out = tk.Label(master, text="Nie wybrano folderu")
        self.label_out.pack()

        tk.Button(master, text="▶ Start", command=self.run).pack(pady=20)

    def choose_input(self):
        folder = filedialog.askdirectory(title="Wybierz folder z wynikami surowymi (XML)")
        if folder:
            self.folder_in = folder
            self.label_in.config(text=folder)

    def choose_template(self):
        path = filedialog.askopenfilename(
            title="Wybierz plik szablonu (xlsx)",
            filetypes=[("Pliki Excel", "*.xlsx")]
        )
        if path:
            self.template_path = path
            self.label_template.config(text=path)

    def choose_output(self):
        folder = filedialog.askdirectory(title="Wybierz folder do zapisania raportu")
        if folder:
            self.folder_out = folder
            self.label_out.config(text=folder)

    def run(self):
        if not self.folder_in or not self.template_path or not self.folder_out:
            messagebox.showerror("Błąd", "Musisz wybrać wszystkie ścieżki!")
            return

        # Szukamy wszystkich plików XML
        pliki = glob.glob(os.path.join(self.folder_in, "*.xml"))
        pliki = sorted(pliki, key=os.path.getmtime)

        if not pliki:
            messagebox.showerror("Błąd", " Nie znaleziono żadnych plików XML w folderze wejściowym.")
            return

        sciezka_raportu = os.path.join(self.folder_out, "raport_wyniki.xlsx")
        shutil.copyfile(self.template_path, sciezka_raportu)

        wb_dst = load_workbook(sciezka_raportu)
        ws_dst = wb_dst.active

        mapa = {
            "id myszki :)" : 1,
            "WBC": 2,
            "LYM%": 3,
            "MON%": 4,
            "GRA%": 5,
            "EOS%": 6,
            "LYM#": 7,
            "MON#": 8,
            "GRA#": 9,
            "EOS#": 10,
            "RBC": 11,
            "HGB": 12,
            "HCT": 13,
            "MCV": 14,
            "MCH": 15,
            "MCHC": 16,
            "RDW": 17,
            "PLT": 18,
            "MPV": 19,
        }

        for i, plik in enumerate(pliki, start=4):
            tree = ET.parse(plik)
            root = tree.getroot()

            data = os.path.getctime(plik)
            for param, col in mapa.items():
                elem = root.find(f".//o[@n='{param}']/d[@n='Value']")
                if elem is not None:
                    wartosc = elem.text
                    if wartosc:
                        # Zamiana przecinka na kropkę, aby float działał poprawnie
                        wartosc = wartosc.replace(",", ".")
                        try:
                            wartosc = float(wartosc)  # konwersja na liczbę
                        except ValueError:
                            pass  # jeśli nie da się zamienić na float, zostaje tekst
                    ws_dst.cell(row=i, column=col, value=wartosc)

            ws_dst.cell(row=i, column=20, value=os.path.basename(plik))

        wb_dst.save(sciezka_raportu)
        messagebox.showinfo("Sukces", f"Raport został zapisany jako:\n{sciezka_raportu}")


if __name__ == "__main__":
    root = tk.Tk()
    app = RaportApp(root)
    root.mainloop()
